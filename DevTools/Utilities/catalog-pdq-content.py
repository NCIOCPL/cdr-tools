#!/usr/bin/env python3

"""Create a spreadsheet listing all of the PDQ Summries in the Drupal CMS.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from requests import get
from sys import stderr

HOST = "www-cms.cancer.gov"
COLS = "Node ID", "CDR ID", "Language", "Title", "URL"
WIDTHS = dict(A=8, B=9, C=10, D=150, E=75)
TYPES = dict(
    pdq_cancer_information_summary="cis",
    pdq_drug_information_summary="dis",
)

auth = "PDQ", open("pdqpw").read().strip()
url = f"https://{HOST}/pdq/api/list?_format=json"
response = get(url, auth=auth)
nodes = {}
for values in response.json():
    nodes[int(values["nid"])] = TYPES[values["type"]]
book = Workbook()
sheet = book.active
sheet.freeze_panes = sheet["A2"]
bold = Font(bold=True)
center = Alignment(horizontal="center")
for col in WIDTHS:
    sheet.column_dimensions[col].width = WIDTHS[col]
    sheet[f"{col}1"].font = bold
    sheet[f"{col}1"].alignment = center
for i, col in enumerate(COLS):
    sheet.cell(column=i+1, row=1, value=col)
row = 2
done = 0
for nid in sorted(nodes):
    url = f"https://{HOST}/pdq/api/{nodes[nid]}/{nid}"
    response = get(url, auth=auth)
    values = response.json()
    if nodes[nid] == "dis":
        sheet.cell(column=1, row=row, value=nid)
        sheet.cell(column=2, row=row, value=int(values["cdr_id"]))
        sheet.cell(column=3, row=row, value="English")
        sheet.cell(column=4, row=row, value=values["title"])
        sheet.cell(column=5, row=row, value=values["url"])
        row += 1
    else:
        sheet.cell(column=1, row=row, value=nid)
        sheet.cell(column=2, row=row, value=int(values["en"]["cdr_id"]))
        sheet.cell(column=3, row=row, value="English")
        sheet.cell(column=4, row=row, value=values["en"]["title"])
        sheet.cell(column=5, row=row, value=values["en"]["url"])
        row += 1
        if "es" in values and "url" in values["es"]:
            sheet.cell(column=1, row=row, value=nid)
            sheet.cell(column=2, row=row, value=int(values["es"]["cdr_id"]))
            sheet.cell(column=3, row=row, value="Spanish")
            sheet.cell(column=4, row=row, value=values["es"]["title"])
            sheet.cell(column=5, row=row, value=values["es"]["url"])
            row += 1
    done += 1
    stderr.write(f"\rfetched {done} of {len(nodes)} nodes")
book.save(filename="pdq-content.xlsx")
stderr.write("\nsaved pdq-content.xslx\n")
